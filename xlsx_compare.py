"""
Compare two Excel files sheet-by-sheet and cell-by-cell and export the differences.

first install dependencies:
pip install pandas openpyxl
v0.1.1
"""

import argparse

import pandas as pd
from openpyxl import load_workbook


# ANSI escape sequences for console coloring
class Colors:  # pylint: disable=too-few-public-methods
    """ANSI color escape sequences"""

    RESET = "\033[0m"
    BOLD = "\033[1m"
    GREEN = "\033[32m"
    YELLOW = "\033[33m"
    RED = "\033[31m"
    BLUE = "\033[34m"
    CYAN = "\033[36m"


def compare_dataframes_cell_by_cell(df_lft, df_rgt, sheet_handle: str) -> pd.DataFrame:
    """
    Compares two dataframes cell-by-cell and returns a dataframe with differences.
    """

    # Normalize data for consistent comparison
    df_lft = (
        df_lft.fillna("")
        .astype(str)
        .apply(lambda col: col.str.strip() if col.dtype == "object" else col)
    )
    df_rgt = (
        df_rgt.fillna("")
        .astype(str)
        .apply(lambda col: col.str.strip() if col.dtype == "object" else col)
    )
    # debug info
    print(f"Shape of File1 ({sheet_handle}): {df_lft.shape}")
    print(f"Shape of File2 ({sheet_handle}): {df_rgt.shape}")

    # Ensure both DataFrames have the same shape by padding
    max_rows = max(df_lft.shape[0], df_rgt.shape[0])
    max_cols = max(df_lft.shape[1], df_rgt.shape[1])
    print(f"Size rows: {max_rows} cols: {max_cols}")

    df_lft = df_lft.reindex(index=range(max_rows), columns=range(max_cols)).fillna("")
    df_rgt = df_rgt.reindex(index=range(max_rows), columns=range(max_cols)).fillna("")

    # Create a DataFrame to store differences
    diff = pd.DataFrame(index=range(max_rows), columns=range(max_cols))
    # print("diff DataFrame created")

    diff_count = 0

    for i in range(max_rows):
        for j in range(max_cols):
            value1 = df_lft.iloc[i, j]
            value2 = df_rgt.iloc[i, j]

            # Debugging
            # print(f"Comparing cell ({i}, {j}): '{repr(value1)}' vs. '{repr(value2)}'")

            if value1 != value2:
                # print(f"Difference: ({i}, {j}): '{repr(value1)}' vs. '{repr(value2)}'")
                diff.iloc[i, j] = f"{value1} -> {value2}"
                diff_count = diff_count + 1
    if diff_count > 0:
        print(f"{Colors.RED}{diff_count} difference found{Colors.RESET}")

    return diff


def get_args():
    """Get file paths from command-line arguments."""

    parser = argparse.ArgumentParser(description="Compare Excel files sheet by sheet.")
    parser.add_argument("file1", help="First Excel file path")
    parser.add_argument("file2", help="Second Excel file path")
    parser.add_argument(
        "output",
        nargs="?",
        default="comparison_output.xlsx",
        help="Output Excel file path",
    )
    parser.add_argument(
        "--sheets",
        help="Comma-separated list of sheet names to compare (default: all sheets)",
    )

    return parser.parse_args()


def save_differences_to_excel(writer: pd.ExcelWriter, sheet_name: str, differences_df: pd.DataFrame) -> None:
    """
    Saves the differences dataframe to the given Excel writer under a safe sheet name.
    """
    safe_sheet_name = sheet_name[:28]  # Truncate for valid Excel sheet name
    # index=False, header=False - not to show additional 1st row and column with indexes
    differences_df.to_excel(writer, sheet_name=f"df-{safe_sheet_name}", index=False, header=False)


def main():
    """Main function to compare two Excel files."""

    args = get_args()

    file1_path = args.file1
    file2_path = args.file2
    # output_path = args.output

    # Processing start
    print(f"File1: {file1_path}")
    print(f"File2: {file2_path}")

    # Load Excel workbooks
    # Using the read_only=True parameter makes openpyxl load the workbook in optimized read-only
    # mode. This is particularly useful for large files because it reduces memory usage and skips
    # features like formatting.
    wb1 = load_workbook(file1_path, read_only=True)
    print("... workbook1 loaded, now loading workbook2")  # to show processing progress
    wb2 = load_workbook(file2_path, read_only=True)

    file1_sheets = wb1.sheetnames
    file2_sheets = wb2.sheetnames

    # Prepare an Excel writer
    print("... Preparing Excel writer")  # to show processing progress
    output_writer = pd.ExcelWriter(args.output, engine="openpyxl")
    print("... Excel writer prepared")  # to show processing progress

    # Create a summary list for the COMPARISON sheet
    comparison_summary = [
        ["File 1", file1_path],
        ["File 2", file2_path],
        [],  # Add an empty row for separation
    ]

    # Compare sheets
    if args.sheets:
        selected_sheets = [s.strip() for s in args.sheets.split(",")]
    else:
        selected_sheets = sorted(set(file1_sheets) | set(file2_sheets))

    for sheet in selected_sheets:
        print(
            f"{Colors.CYAN}Processing sheet: {sheet}{Colors.RESET}"
        )  # Cyan for processing progress
        if sheet not in file1_sheets:
            if sheet not in file2_sheets:
                # Sheet is only in args.sheet
                print("... only in the --sheets list")
            else:
                # Sheet only in file2
                print("... only in file2")
                comparison_summary.append([sheet, "Only in file2"])
        elif sheet not in file2_sheets:
            # Sheet only in file1
            print("... only in file1")
            comparison_summary.append([sheet, "Only in file1"])
        else:
            # Sheet in both files
            # dtype=str : all data is treated as strings and any non-standard formats are handled
            # header=None : the first row is also compared
            df1 = pd.read_excel(file1_path, sheet_name=sheet, dtype=str, header=None)
            df2 = pd.read_excel(file2_path, sheet_name=sheet, dtype=str, header=None)
            # debug
            # print(f"Data from File1 ({sheet}):")
            # print(df1)
            # print(f"Data from File2 ({sheet}):")
            # print(df2)

            # Compare content cell by cell
            differences = compare_dataframes_cell_by_cell(df1, df2, sheet)
            if differences.isnull().all().all():
                comparison_summary.append([sheet, "No differences"])
                print("... no difference")
            else:
                print(f"{Colors.RED}... some difference{Colors.RESET}")
                # Save differences to a separate sheet
                save_differences_to_excel(output_writer, sheet, differences)
                comparison_summary.append([sheet, "Differences found"])

    # Write the summary to the COMPARISON sheet
    comparison_df = pd.DataFrame(
        comparison_summary,
        columns=["Sheet Name", "Status"],  # Header row for the comparison summary
    )
    comparison_df.to_excel(output_writer, sheet_name="COMPARISON", index=False)

    # Save the output file
    output_writer.close()
    print(f"Comparison completed. Output saved as '{args.output}'.")


if __name__ == "__main__":
    main()
